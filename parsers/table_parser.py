"""
============================================================
FILE NAME
table_parser.py

PURPOSE
Parse Excel (XLSX) / CSV tables into tab-separated text.

INPUT
File path

OUTPUT
{"text": ..., "source": ..., "type": "table"}

DEPENDENCIES
openpyxl, csv

LAST UPDATED
2026-08-02
============================================================
"""

import csv
from pathlib import Path


class TableParser:

    def parse(self, source):

        path = Path(str(source))

        suffix = path.suffix.lower()

        if suffix == ".csv":
            text = self._parse_csv(path)
        else:
            text = self._parse_xlsx(path)

        return {
            "text": text,
            "source": str(path),
            "type": "table",
            "extension": suffix
        }

    def _parse_xlsx(self, path):
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                str(path),
                read_only=True,
                data_only=True
            )

            lines = []

            for sheet in workbook.worksheets:

                lines.append(f"[Sheet: {sheet.title}]")

                for row in sheet.iter_rows(values_only=True):
                    cells = [
                        str(cell)
                        for cell in row
                        if cell is not None
                    ]
                    if cells:
                        lines.append("\t".join(cells))

            workbook.close()

            return "\n".join(lines)
        except Exception:
            return ""

    def _parse_csv(self, path):
        try:
            lines = []

            with open(path, newline="", encoding="utf-8-sig") as handle:
                reader = csv.reader(handle)

                for row in reader:
                    lines.append("\t".join(row))

            return "\n".join(lines)
        except Exception:
            return ""
